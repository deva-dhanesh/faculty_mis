"""
bulk_utils.py
Utilities for generating downloadable Excel templates and
parsing bulk-uploaded CSV / Excel files for publications & projects.
"""

import io
import csv
from datetime import datetime, date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ── colour palette ──────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1E293B")   # dark navy
SAMPLE_FILL= PatternFill("solid", fgColor="F1F5F9")   # light grey
HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
SAMPLE_FONT= Font(color="64748B", size=9, italic=True)
BODY_FONT  = Font(size=9)
THIN       = Side(border_style="thin", color="CBD5E1")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER     = Alignment(horizontal="center", vertical="center")
LEFT       = Alignment(horizontal="left",  vertical="center", wrap_text=True)


def _style_header(ws, row, cols):
    for col_idx, (header, width) in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 22


def _style_sample(ws, row, values):
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font      = SAMPLE_FONT
        cell.fill      = SAMPLE_FILL
        cell.alignment = LEFT
        cell.border    = BORDER
    ws.row_dimensions[row].height = 18


def _add_dropdown(ws, col_letter, values, start_row=3, end_row=1000):
    formula = '"' + ','.join(values) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.add_data_validation(dv)


# ============================================================
# BUILD PUBLICATIONS TEMPLATE
# ============================================================

PUB_COLS = [
    ("Title *",                                 40),
    ("Author Position",                         20),
    ("Scopus/ORCID/Scholar ID",                 28),
    ("Name of Journal",                         30),
    ("Publication Date (DD-MM-YYYY)",            22),
    ("ISSN / ISBN",                              16),
    ("h-index",                                  10),
    ("Citation Index",                           14),
    ("Journal Quartile (Q1/Q2/Q3/Q4/None)",      18),
    ("Type (Journal/Conference)",                 20),
    ("Impact Factor",                            14),
    ("Indexing",                                 18),
    ("DOI",                                      34),
    ("Article Link (URL)",                       36),
]

PUB_SAMPLE = [
    "Nanometre-scale thermometry in a living cell",
    "First Author",
    "0000-0002-1234-5678",
    "Nature",
    "15-03-2024",
    "1234-5678",
    "12",
    "150",
    "Q1",
    "Journal",
    "4.52",
    "Scopus",
    "10.1038/nature12373",
    "https://doi.org/10.1038/nature12373",
]

AUTHOR_POSITION_OPTIONS = ["First Author", "Second Author", "Third Author",
                           "Corresponding Author", "Co-Author", "Custom"]
QUARTILE_OPTIONS  = ["Q1", "Q2", "Q3", "Q4", "None"]
PUB_TYPE_OPTIONS  = ["Journal", "Conference"]
INDEXING_OPTIONS   = ["Scopus", "WoS", "UGC CARE", "PubMed", "Others"]


def build_publications_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Publications"
    ws.freeze_panes = "A3"

    num_cols = len(PUB_COLS)
    last_col = get_column_letter(num_cols)

    # Title row
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value     = "Faculty MIS — Bulk Publications Upload Template"
    title_cell.font      = Font(bold=True, size=12, color="0F172A")
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 26

    # Header row
    _style_header(ws, 2, PUB_COLS)

    # Sample row
    _style_sample(ws, 3, PUB_SAMPLE)

    # Note row
    ws.merge_cells(f"A4:{last_col}4")
    note = ws["A4"]
    note.value     = "⚠  Delete the sample row (row 3) before uploading. * = required. Dates must be DD-MM-YYYY."
    note.font      = Font(size=8, color="94A3B8", italic=True)
    note.alignment = LEFT

    # Dropdowns
    _add_dropdown(ws, "B", AUTHOR_POSITION_OPTIONS, start_row=5)      # Author Position
    _add_dropdown(ws, "I", QUARTILE_OPTIONS,        start_row=5)      # Journal Quartile
    _add_dropdown(ws, "J", PUB_TYPE_OPTIONS,        start_row=5)      # Type
    _add_dropdown(ws, "L", INDEXING_OPTIONS,         start_row=5)      # Indexing

    for r in range(5, 1001):
        ws.row_dimensions[r].height = 16
        for c in range(1, num_cols + 1):
            ws.cell(r, c).font   = BODY_FONT
            ws.cell(r, c).border = BORDER

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# BUILD PROJECTS TEMPLATE
# ============================================================

PROJ_COLS = [
    ("Scheme / Project Name *",            45),
    ("PI / Co-PI",                         28),
    ("Funding Agency",                     24),
    ("Type (Govt./Non-Govt.)",             20),
    ("Department",                         22),
    ("Date of Award (YYYY-MM-DD)",         22),
    ("Total Fund Granted (INR Lakhs)",     24),
    ("Duration (Years)",                   16),
    ("Project Status",                     18),
    ("Date of Completion (YYYY-MM-DD)",    24),
    ("Objectives / Goals",                 36),
    ("Collaborating Institutions",         32),
    ("Sanction Order Link (URL)",          36),
]

PROJ_SAMPLE = [
    "AI-driven crop disease detection",
    "Dr. A. Kumar / Dr. B. Rani",
    "DST",
    "Govt.",
    "Computer Science",
    "2024-01-15",
    5.50,
    "2",
    "Ongoing",
    "",
    "To develop AI models for early disease detection in crops.",
    "IIT Madras",
    "https://drive.google.com/sample",
]

PROJ_TYPE_OPTIONS   = ["Govt.", "Non-Govt."]
PROJ_STATUS_OPTIONS = ["Ongoing", "Completed"]


def build_projects_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projects"
    ws.freeze_panes = "A3"

    num_cols = len(PROJ_COLS)
    last_col = get_column_letter(num_cols)

    # Title row
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value     = "Faculty MIS — Bulk Research Projects Upload Template"
    title_cell.font      = Font(bold=True, size=12, color="0F172A")
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 26

    # Header row
    _style_header(ws, 2, PROJ_COLS)

    # Sample row
    _style_sample(ws, 3, PROJ_SAMPLE)

    # Note row
    ws.merge_cells(f"A4:{last_col}4")
    note = ws["A4"]
    note.value     = "⚠  Delete the sample row (row 3) before uploading. * = required. Dates must be YYYY-MM-DD."
    note.font      = Font(size=8, color="94A3B8", italic=True)
    note.alignment = LEFT

    # Dropdowns
    _add_dropdown(ws, "D", PROJ_TYPE_OPTIONS,   start_row=5)
    _add_dropdown(ws, "I", PROJ_STATUS_OPTIONS,  start_row=5)

    for r in range(5, 1001):
        ws.row_dimensions[r].height = 16
        for c in range(1, num_cols + 1):
            ws.cell(r, c).font   = BODY_FONT
            ws.cell(r, c).border = BORDER

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# PARSE UPLOADED FILE  →  list of dicts
# ============================================================

def _read_workbook_rows(file_bytes: bytes, sheet_index: int = 0):
    """Return list of row-tuples from an xlsx/xls file (skips header row)."""
    wb   = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws   = wb.worksheets[sheet_index]
    rows = list(ws.iter_rows(values_only=True))
    # Skip title row (row 1) and header row (row 2)
    # If first non-empty row is a header (contains "Title" etc.) skip it
    data_rows = []
    for row in rows[1:]:          # skip merged title in row 1
        if row[0] is None:
            continue
        if str(row[0]).strip().lower() in ("title", "title *", ""):
            continue              # skip header row
        data_rows.append(row)
    return data_rows


def _read_csv_rows(file_bytes: bytes):
    """Return list of row-dicts from a CSV file."""
    text    = file_bytes.decode("utf-8-sig", errors="replace")
    reader  = csv.DictReader(io.StringIO(text))
    return list(reader)


def _safe_int(val):
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


def _safe_float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _safe_date(val):
    if not val:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d-%b-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(val).strip(), fmt).date()
        except ValueError:
            pass
    return None


# ── Publications ─────────────────────────────────────────────────────────────

def parse_publications(file_bytes: bytes, filename: str):
    """
    Returns (records, errors)
    records : list of dicts ready to create FacultyPublication objects
    errors  : list of (row_num, message) tuples
    """
    records = []
    errors  = []
    is_csv  = filename.lower().endswith(".csv")

    if is_csv:
        raw = _read_csv_rows(file_bytes)
        for i, row in enumerate(raw, start=2):
            title = (row.get("Title *") or row.get("Title") or "").strip()
            if not title:
                errors.append((i, "Title is required"))
                continue
            records.append({
                "title"            : title,
                "author_position"  : (row.get("Author Position") or "").strip() or None,
                "scholar_id"       : (row.get("Scopus/ORCID/Scholar ID") or row.get("Scholar ID") or "").strip() or None,
                "journal"          : (row.get("Name of Journal") or row.get("Journal") or "").strip() or None,
                "publication_date" : _safe_date(row.get("Publication Date (DD-MM-YYYY)") or row.get("Publication Date")),
                "issn_isbn"        : (row.get("ISSN / ISBN") or row.get("ISSN/ISBN") or "").strip() or None,
                "h_index"          : (row.get("h-index") or "").strip() or None,
                "citation_index"   : (row.get("Citation Index") or "").strip() or None,
                "journal_quartile" : (row.get("Journal Quartile (Q1/Q2/Q3/Q4)") or row.get("Quartile") or "").strip() or None,
                "publication_type" : (row.get("Type (Journal/Conference)") or row.get("Type") or "").strip() or None,
                "impact_factor"    : (row.get("Impact Factor") or "").strip() or None,
                "indexing"         : (row.get("Indexing") or "").strip() or None,
                "doi"              : (row.get("DOI") or "").strip() or None,
                "article_link"     : (row.get("Article Link (URL)") or row.get("Article Link") or "").strip() or None,
            })
    else:
        raw = _read_workbook_rows(file_bytes)
        for i, row in enumerate(raw, start=3):
            title = str(row[0]).strip() if row[0] else ""
            if not title:
                errors.append((i, "Title is required"))
                continue
            records.append({
                "title"            : title,
                "author_position"  : str(row[1]).strip() if len(row) > 1 and row[1] else None,
                "scholar_id"       : str(row[2]).strip() if len(row) > 2 and row[2] else None,
                "journal"          : str(row[3]).strip() if len(row) > 3 and row[3] else None,
                "publication_date" : _safe_date(row[4]) if len(row) > 4 else None,
                "issn_isbn"        : str(row[5]).strip() if len(row) > 5 and row[5] else None,
                "h_index"          : str(row[6]).strip() if len(row) > 6 and row[6] else None,
                "citation_index"   : str(row[7]).strip() if len(row) > 7 and row[7] else None,
                "journal_quartile" : str(row[8]).strip() if len(row) > 8 and row[8] else None,
                "publication_type" : str(row[9]).strip() if len(row) > 9 and row[9] else None,
                "impact_factor"    : str(row[10]).strip() if len(row) > 10 and row[10] else None,
                "indexing"         : str(row[11]).strip() if len(row) > 11 and row[11] else None,
                "doi"              : str(row[12]).strip() if len(row) > 12 and row[12] else None,
                "article_link"     : str(row[13]).strip() if len(row) > 13 and row[13] else None,
            })

    return records, errors


# ── Projects ─────────────────────────────────────────────────────────────────

def parse_projects(file_bytes: bytes, filename: str):
    """
    Returns (records, errors)
    """
    records = []
    errors  = []
    is_csv  = filename.lower().endswith(".csv")

    if is_csv:
        raw = _read_csv_rows(file_bytes)
        for i, row in enumerate(raw, start=2):
            scheme_name = (row.get("Scheme / Project Name *") or row.get("Scheme / Project Name") or "").strip()
            if not scheme_name:
                errors.append((i, "Scheme / Project Name is required"))
                continue
            records.append({
                "scheme_name"                : scheme_name,
                "pi_co_pi"                   : (row.get("PI / Co-PI") or "").strip() or None,
                "funding_agency"             : (row.get("Funding Agency") or "").strip() or None,
                "project_type"               : (row.get("Type (Govt./Non-Govt.)") or "").strip() or None,
                "department"                 : (row.get("Department") or "").strip() or None,
                "date_of_award"              : _safe_date(row.get("Date of Award (YYYY-MM-DD)")),
                "amount"                     : _safe_float(row.get("Total Fund Granted (INR Lakhs)")),
                "duration_years"             : (row.get("Duration (Years)") or "").strip() or None,
                "status"                     : (row.get("Project Status") or "Ongoing").strip(),
                "date_of_completion"         : _safe_date(row.get("Date of Completion (YYYY-MM-DD)")),
                "objectives"                 : (row.get("Objectives / Goals") or "").strip() or None,
                "collaborating_institutions" : (row.get("Collaborating Institutions") or "").strip() or None,
                "document_link"              : (row.get("Sanction Order Link (URL)") or "").strip() or None,
            })
    else:
        raw = _read_workbook_rows(file_bytes)
        for i, row in enumerate(raw, start=3):
            scheme_name = str(row[0]).strip() if row[0] else ""
            if not scheme_name:
                errors.append((i, "Scheme / Project Name is required"))
                continue
            records.append({
                "scheme_name"                : scheme_name,
                "pi_co_pi"                   : str(row[1]).strip() if row[1] else None,
                "funding_agency"             : str(row[2]).strip() if row[2] else None,
                "project_type"               : str(row[3]).strip() if row[3] else None,
                "department"                 : str(row[4]).strip() if row[4] else None,
                "date_of_award"              : _safe_date(row[5]),
                "amount"                     : _safe_float(row[6]),
                "duration_years"             : str(row[7]).strip() if row[7] else None,
                "status"                     : str(row[8]).strip() if row[8] else "Ongoing",
                "date_of_completion"         : _safe_date(row[9]),
                "objectives"                 : str(row[10]).strip() if row[10] else None,
                "collaborating_institutions" : str(row[11]).strip() if row[11] else None,
                "document_link"              : str(row[12]).strip() if row[12] else None,
            })

    return records, errors


# ============================================================
# BUILD FACULTY USERS TEMPLATE  (admin bulk-create accounts)
# ============================================================

FACULTY_USER_COLS = [
    ("Employee ID *",               18),
    ("Full Name *",                 30),
    ("Email *",                     32),
    ("Phone *",                     16),
    ("Department",                  26),
    ("Designation",                 26),
    ("Password *",                  22),
]

FACULTY_USER_SAMPLE = [
    "EMP001", "Dr. Jane Smith", "jane.smith@university.edu",
    "9876543210", "Computer Science", "Assistant Professor", "Temp@1234",
]

DESIGNATION_OPTIONS = [
    "Professor", "Associate Professor", "Assistant Professor",
    "Senior Lecturer", "Lecturer", "Guest Faculty"
]


def build_faculty_users_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Faculty Users"
    ws.freeze_panes = "A3"
    n = len(FACULTY_USER_COLS)
    last = get_column_letter(n)
    ws.merge_cells(f"A1:{last}1")
    t = ws["A1"]
    t.value     = "Faculty MIS — Bulk Faculty Accounts Upload Template"
    t.font      = Font(bold=True, size=12, color="0F172A")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 26
    _style_header(ws, 2, FACULTY_USER_COLS)
    _style_sample(ws, 3, FACULTY_USER_SAMPLE)
    ws.merge_cells(f"A4:{last}4")
    note = ws["A4"]
    note.value     = "⚠  Delete the sample row (row 3) before uploading. * = required. Faculty will be forced to reset password on first login."
    note.font      = Font(size=8, color="94A3B8", italic=True)
    note.alignment = LEFT
    _add_dropdown(ws, "F", DESIGNATION_OPTIONS, start_row=5)
    for r in range(5, 1001):
        ws.row_dimensions[r].height = 16
        for c in range(1, n + 1):
            ws.cell(r, c).font   = BODY_FONT
            ws.cell(r, c).border = BORDER
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# BUILD FACULTY PROFILES TEMPLATE  (admin bulk-upsert profiles)
# ============================================================

FACULTY_PROFILE_COLS = [
    ("Employee ID *",                18),
    ("Full Name",                    30),
    ("PAN",                          14),
    ("Designation",                  26),
    ("Date of Joining (YYYY-MM-DD)", 24),
    ("Date of Birth (YYYY-MM-DD)",   22),
    ("Appointment Nature",           22),
    ("Qualification",                28),
    ("Department",                   26),
    ("Experience Years",             16),
    ("Mobile",                       16),
    ("Personal Email",               30),
    ("University Email",             30),
    ("Specialization",               32),
]

FACULTY_PROFILE_SAMPLE = [
    "EMP001", "Dr. Jane Smith", "ABCDE1234F", "Assistant Professor",
    "2020-07-15", "1985-03-22", "Permanent", "Ph.D. Computer Science",
    "Computer Science", 5, "9876543210",
    "jane.personal@gmail.com", "jane.smith@university.edu", "Machine Learning, NLP",
]

APPOINTMENT_NATURE_OPTIONS = ["Permanent", "Temporary", "Guest", "Contract", "Ad-hoc"]


def build_faculty_profiles_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Faculty Profiles"
    ws.freeze_panes = "A3"
    n = len(FACULTY_PROFILE_COLS)
    last = get_column_letter(n)
    ws.merge_cells(f"A1:{last}1")
    t = ws["A1"]
    t.value     = "Faculty MIS — Bulk Faculty Profiles Upload Template"
    t.font      = Font(bold=True, size=12, color="0F172A")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 26
    _style_header(ws, 2, FACULTY_PROFILE_COLS)
    _style_sample(ws, 3, FACULTY_PROFILE_SAMPLE)
    ws.merge_cells(f"A4:{last}4")
    note = ws["A4"]
    note.value     = "⚠  Delete the sample row (row 3) before uploading. * = required. Employee ID must match an existing faculty account. Dates must be YYYY-MM-DD."
    note.font      = Font(size=8, color="94A3B8", italic=True)
    note.alignment = LEFT
    _add_dropdown(ws, "D", DESIGNATION_OPTIONS,        start_row=5)
    _add_dropdown(ws, "G", APPOINTMENT_NATURE_OPTIONS, start_row=5)
    for r in range(5, 1001):
        ws.row_dimensions[r].height = 16
        for c in range(1, n + 1):
            ws.cell(r, c).font   = BODY_FONT
            ws.cell(r, c).border = BORDER
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================# BUILD COURSES ATTENDED TEMPLATE
# ============================================================

CA_COLS = [
    ("Course Name *",                      40),
    ("Online Course / Platform",           36),
    ("Date From (YYYY-MM-DD)",             20),
    ("Date To   (YYYY-MM-DD)",             20),
    ("Mode",                               28),
    ("Contact Hours",                      14),
    ("Offered By",                         28),
    ("Certificate Link",                   36),
]

CA_SAMPLE = [
    "Python for Data Science",
    "Machine Learning — Coursera (Andrew Ng)",
    "2024-01-10",
    "2024-03-10",
    "Online — Coursera",
    40,
    "Coursera / Stanford University",
    "https://coursera.org/verify/abc123",
]

MODE_ATTENDED_OPTIONS = [
    "Offline — Institution",
    "Online — Coursera",
    "Online — LinkedIn Learning",
    "Online — NPTEL",
    "Online — SWAYAM",
    "Online — MOOCs",
    "Online — e-Pathshala",
    "Other",
]


def build_courses_attended_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Courses Attended"
    ws.freeze_panes = "A3"
    n    = len(CA_COLS)
    last = get_column_letter(n)

    ws.merge_cells(f"A1:{last}1")
    t        = ws["A1"]
    t.value  = "Faculty MIS — Bulk Courses Attended Upload Template"
    t.font   = Font(bold=True, size=12, color="0F172A")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 26

    _style_header(ws, 2, CA_COLS)
    _style_sample(ws, 3, CA_SAMPLE)

    ws.merge_cells(f"A4:{last}4")
    note        = ws["A4"]
    note.value  = "⚠  Delete the sample row (row 3) before uploading. * = required. Dates must be YYYY-MM-DD."
    note.font   = Font(size=8, color="94A3B8", italic=True)
    note.alignment = LEFT

    _add_dropdown(ws, "E", MODE_ATTENDED_OPTIONS, start_row=5)

    for r in range(5, 1001):
        ws.row_dimensions[r].height = 16
        for c in range(1, n + 1):
            ws.cell(r, c).font   = BODY_FONT
            ws.cell(r, c).border = BORDER

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# BUILD COURSES OFFERED TEMPLATE
# ============================================================

CO_COLS = [
    ("Course Name *",                      40),
    ("Online Course / Platform",           36),
    ("Credits Assigned",                   22),
    ("Program Name",                       26),
    ("Department",                         24),
    ("Date From (YYYY-MM-DD)",             20),
    ("Date To   (YYYY-MM-DD)",             20),
    ("Times Offered",                      14),
    ("Mode",                               28),
    ("Contact Hours",                      14),
    ("Students Enrolled Link",             36),
    ("Students Completing Link",           36),
    ("Attendance Link",                    36),
    ("Brochure / Syllabus Link",           36),
    ("Certificate Link",                   36),
]

CO_SAMPLE = [
    "Advanced Python Programming",
    "Deep Learning Specialisation — Coursera",
    "Yes — 2 Credits",
    "B.Tech CSE",
    "Computer Science",
    "2024-07-01",
    "2024-11-30",
    1,
    "Offline — Institution",
    45,
    "https://drive.google.com/enrolled",
    "https://drive.google.com/completing",
    "https://drive.google.com/attendance",
    "https://drive.google.com/brochure",
    "https://drive.google.com/certificate",
]

MODE_OFFERED_OPTIONS = [
    "Offline — Institution",
    "Online — Coursera",
    "Online — LinkedIn Learning",
    "Online — NPTEL",
    "Online — SWAYAM",
    "Online — MOOCs",
    "Online — e-Pathshala",
    "Blended",
    "Other",
]


def build_courses_offered_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Courses Offered"
    ws.freeze_panes = "A3"
    n    = len(CO_COLS)
    last = get_column_letter(n)

    ws.merge_cells(f"A1:{last}1")
    t        = ws["A1"]
    t.value  = "Faculty MIS — Bulk Courses Offered Upload Template"
    t.font   = Font(bold=True, size=12, color="0F172A")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 26

    _style_header(ws, 2, CO_COLS)
    _style_sample(ws, 3, CO_SAMPLE)

    ws.merge_cells(f"A4:{last}4")
    note        = ws["A4"]
    note.value  = "⚠  Delete the sample row (row 3) before uploading. * = required. Dates must be YYYY-MM-DD."
    note.font   = Font(size=8, color="94A3B8", italic=True)
    note.alignment = LEFT

    _add_dropdown(ws, "I", MODE_OFFERED_OPTIONS, start_row=5)

    for r in range(5, 1001):
        ws.row_dimensions[r].height = 16
        for c in range(1, n + 1):
            ws.cell(r, c).font   = BODY_FONT
            ws.cell(r, c).border = BORDER

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# PARSE COURSES ATTENDED
# ============================================================

def parse_courses_attended(file_bytes: bytes, filename: str):
    """Returns (records, errors)."""
    records, errors = [], []
    is_csv = filename.lower().endswith(".csv")

    if is_csv:
        raw = _read_csv_rows(file_bytes)
        for i, row in enumerate(raw, start=2):
            name = (row.get("Course Name *") or row.get("Course Name") or "").strip()
            if not name:
                errors.append((i, "Course Name is required"))
                continue
            records.append({
                "course_name":       name,
                "online_course_name": (row.get("Online Course / Platform") or "").strip() or None,
                "date_from":         _safe_date(row.get("Date From (YYYY-MM-DD)") or row.get("Date From")),
                "date_to":           _safe_date(row.get("Date To   (YYYY-MM-DD)") or row.get("Date To")),
                "mode":              (row.get("Mode") or "").strip() or None,
                "contact_hours":     _safe_float(row.get("Contact Hours")),
                "offered_by":        (row.get("Offered By") or "").strip() or None,
                "certificate_link":  (row.get("Certificate Link") or "").strip() or None,
            })
    else:
        raw = _read_workbook_rows(file_bytes)
        for i, row in enumerate(raw, start=3):
            name = str(row[0]).strip() if row[0] else ""
            if not name:
                errors.append((i, "Course Name is required"))
                continue
            records.append({
                "course_name":        name,
                "online_course_name": str(row[1]).strip() if row[1] else None,
                "date_from":          _safe_date(row[2]),
                "date_to":            _safe_date(row[3]),
                "mode":               str(row[4]).strip() if row[4] else None,
                "contact_hours":      _safe_float(row[5]),
                "offered_by":         str(row[6]).strip() if row[6] else None,
                "certificate_link":   str(row[7]).strip() if row[7] else None,
            })

    return records, errors


# ============================================================
# PARSE COURSES OFFERED
# ============================================================

def parse_courses_offered(file_bytes: bytes, filename: str):
    """Returns (records, errors)."""
    records, errors = [], []
    is_csv = filename.lower().endswith(".csv")

    if is_csv:
        raw = _read_csv_rows(file_bytes)
        for i, row in enumerate(raw, start=2):
            name = (row.get("Course Name *") or row.get("Course Name") or "").strip()
            if not name:
                errors.append((i, "Course Name is required"))
                continue
            records.append({
                "course_name":            name,
                "online_course_name":     (row.get("Online Course / Platform") or "").strip() or None,
                "credits_assigned":       (row.get("Credits Assigned") or "").strip() or None,
                "program_name":           (row.get("Program Name") or "").strip() or None,
                "department":             (row.get("Department") or "").strip() or None,
                "date_from":              _safe_date(row.get("Date From (YYYY-MM-DD)") or row.get("Date From")),
                "date_to":                _safe_date(row.get("Date To   (YYYY-MM-DD)") or row.get("Date To")),
                "times_offered":          _safe_int(row.get("Times Offered")),
                "mode":                   (row.get("Mode") or "").strip() or None,
                "contact_hours":          _safe_float(row.get("Contact Hours")),
                "students_enrolled_link": (row.get("Students Enrolled Link") or "").strip() or None,
                "students_completing_link": (row.get("Students Completing Link") or "").strip() or None,
                "attendance_link":        (row.get("Attendance Link") or "").strip() or None,
                "brochure_link":          (row.get("Brochure / Syllabus Link") or "").strip() or None,
                "certificate_link":       (row.get("Certificate Link") or "").strip() or None,
            })
    else:
        raw = _read_workbook_rows(file_bytes)
        for i, row in enumerate(raw, start=3):
            name = str(row[0]).strip() if row[0] else ""
            if not name:
                errors.append((i, "Course Name is required"))
                continue
            records.append({
                "course_name":            name,
                "online_course_name":     str(row[1]).strip()  if row[1]  else None,
                "credits_assigned":       str(row[2]).strip()  if row[2]  else None,
                "program_name":           str(row[3]).strip()  if row[3]  else None,
                "department":             str(row[4]).strip()  if row[4]  else None,
                "date_from":              _safe_date(row[5]),
                "date_to":                _safe_date(row[6]),
                "times_offered":          _safe_int(row[7]),
                "mode":                   str(row[8]).strip()  if row[8]  else None,
                "contact_hours":          _safe_float(row[9]),
                "students_enrolled_link": str(row[10]).strip() if row[10] else None,
                "students_completing_link": str(row[11]).strip() if row[11] else None,
                "attendance_link":        str(row[12]).strip() if row[12] else None,
                "brochure_link":          str(row[13]).strip() if row[13] else None,
                "certificate_link":       str(row[14]).strip() if row[14] else None,
            })

    return records, errors


# ============================================================
# BUILD AWARDS TEMPLATE
# ============================================================

AWARD_COLS = [
    ("Title of Innovation / Academic Achievement *",   50),
    ("Nature of Award",                                30),
    ("Event Level",                                    20),
    ("Date of Award (YYYY-MM-DD)",                     22),
    ("Category",                                       26),
    ("Name of Awarding Agency",                        36),
    ("Award Amount (if any)",                          20),
    ("Project / Research Area",                        30),
    ("Collaborators (if any)",                         30),
    ("Link to Relevant Documents",                     40),
]

AWARD_SAMPLE = [
    "Best Research Paper Award — IEEE 2024",
    "Research",
    "National",
    "2024-03-15",
    "Teachers",
    "IEEE India Council",
    "₹10,000",
    "Machine Learning in Healthcare",
    "Dr. A. Kumar, Dr. B. Singh",
    "https://drive.google.com/award-certificate",
]

AWARD_NATURE_OPTIONS = [
    "Academic",
    "Research",
    "Innovations",
    "Sports",
    "Cultural",
    "Alumni",
    "Women Cell",
]

AWARD_LEVEL_OPTIONS = [
    "Local",
    "State",
    "National",
    "International",
]

AWARD_CATEGORY_OPTIONS = [
    "Department",
    "Teachers",
    "Research Scholars",
]


def build_awards_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Awards"
    ws.freeze_panes = "A3"
    n    = len(AWARD_COLS)
    last = get_column_letter(n)

    ws.merge_cells(f"A1:{last}1")
    t        = ws["A1"]
    t.value  = "Faculty MIS — Bulk Awards Upload Template"
    t.font   = Font(bold=True, size=12, color="0F172A")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 26

    _style_header(ws, 2, AWARD_COLS)
    _style_sample(ws, 3, AWARD_SAMPLE)

    ws.merge_cells(f"A4:{last}4")
    note        = ws["A4"]
    note.value  = "⚠  Delete the sample row (row 3) before uploading. * = required. Dates must be YYYY-MM-DD."
    note.font   = Font(size=8, color="94A3B8", italic=True)
    note.alignment = LEFT

    _add_dropdown(ws, "B", AWARD_NATURE_OPTIONS, start_row=5)
    _add_dropdown(ws, "C", AWARD_LEVEL_OPTIONS, start_row=5)
    _add_dropdown(ws, "E", AWARD_CATEGORY_OPTIONS, start_row=5)

    for r in range(5, 1001):
        ws.row_dimensions[r].height = 16
        for c in range(1, n + 1):
            ws.cell(r, c).font   = BODY_FONT
            ws.cell(r, c).border = BORDER

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# PARSE AWARDS
# ============================================================

def parse_awards(file_bytes: bytes, filename: str):
    """Returns (records, errors)."""
    records, errors = [], []
    is_csv = filename.lower().endswith(".csv")

    if is_csv:
        raw = _read_csv_rows(file_bytes)
        for i, row in enumerate(raw, start=2):
            title = (row.get("Title of Innovation / Academic Achievement *") or
                     row.get("Title of Innovation / Academic Achievement") or "").strip()
            if not title:
                errors.append((i, "Title is required"))
                continue
            records.append({
                "title":           title,
                "nature_of_award": (row.get("Nature of Award") or "").strip() or None,
                "event_level":     (row.get("Event Level") or "").strip() or None,
                "date_of_award":   _safe_date(row.get("Date of Award (YYYY-MM-DD)") or row.get("Date of Award")),
                "category":        (row.get("Category") or "").strip() or None,
                "awarding_agency": (row.get("Name of Awarding Agency") or "").strip() or None,
                "award_amount":    (row.get("Award Amount (if any)") or "").strip() or None,
                "research_area":   (row.get("Project / Research Area") or "").strip() or None,
                "collaborators":   (row.get("Collaborators (if any)") or "").strip() or None,
                "document_link":   (row.get("Link to Relevant Documents") or "").strip() or None,
            })
    else:
        raw = _read_workbook_rows(file_bytes)
        for i, row in enumerate(raw, start=3):
            title = str(row[0]).strip() if row[0] else ""
            if not title:
                errors.append((i, "Title is required"))
                continue
            records.append({
                "title":           title,
                "nature_of_award": str(row[1]).strip() if row[1] else None,
                "event_level":     str(row[2]).strip() if row[2] else None,
                "date_of_award":   _safe_date(row[3]),
                "category":        str(row[4]).strip() if row[4] else None,
                "awarding_agency": str(row[5]).strip() if row[5] else None,
                "award_amount":    str(row[6]).strip() if row[6] else None,
                "research_area":   str(row[7]).strip() if row[7] else None,
                "collaborators":   str(row[8]).strip() if row[8] else None,
                "document_link":   str(row[9]).strip() if row[9] else None,
            })

    return records, errors


# ============================================================# PARSE FACULTY USERS  (admin bulk-create accounts)
# ============================================================

def parse_faculty_users(file_bytes: bytes, filename: str):
    """Returns (records, errors). records are dicts for User creation."""
    records, errors = [], []
    is_csv = filename.lower().endswith(".csv")
    if is_csv:
        raw = _read_csv_rows(file_bytes)
        for i, row in enumerate(raw, start=2):
            eid   = (row.get("Employee ID *") or row.get("Employee ID") or row.get("employee_id") or "").strip()
            name  = (row.get("Full Name *")   or row.get("Full Name")   or row.get("full_name")   or "").strip()
            email = (row.get("Email *")        or row.get("Email")       or row.get("email")        or "").strip().lower()
            phone = (row.get("Phone *")        or row.get("Phone")       or row.get("phone")        or "").strip()
            dept  = (row.get("Department")     or row.get("department")  or "").strip() or None
            desig = (row.get("Designation")    or row.get("designation") or "").strip() or None
            pwd   = (row.get("Password *")     or row.get("Password")    or row.get("password")    or "").strip()
            if not eid or not email or not phone or not pwd:
                errors.append((i, f"Row {i}: Employee ID, Email, Phone and Password are all required"))
                continue
            records.append({"employee_id": eid, "full_name": name, "email": email,
                            "phone": phone, "department": dept, "designation": desig, "password": pwd})
    else:
        raw = _read_workbook_rows(file_bytes)
        for i, row in enumerate(raw, start=3):
            eid   = str(row[0]).strip() if row[0] else ""
            name  = str(row[1]).strip() if row[1] else ""
            email = str(row[2]).strip().lower() if row[2] else ""
            phone = str(row[3]).strip() if row[3] else ""
            dept  = str(row[4]).strip() if row[4] else None
            desig = str(row[5]).strip() if row[5] else None
            pwd   = str(row[6]).strip() if row[6] else ""
            if not eid or not email or not phone or not pwd:
                errors.append((i, f"Row {i}: Employee ID, Email, Phone and Password are all required"))
                continue
            records.append({"employee_id": eid, "full_name": name, "email": email,
                            "phone": phone, "department": dept, "designation": desig, "password": pwd})
    return records, errors


# ============================================================
# PARSE FACULTY PROFILES  (admin bulk-upsert profiles)
# ============================================================

def parse_faculty_profiles(file_bytes: bytes, filename: str):
    """Returns (records, errors). records are dicts for FacultyProfile upsert."""
    records, errors = [], []
    is_csv = filename.lower().endswith(".csv")
    if is_csv:
        raw = _read_csv_rows(file_bytes)
        for i, row in enumerate(raw, start=2):
            eid = (row.get("Employee ID *") or row.get("Employee ID") or row.get("employee_id") or "").strip()
            if not eid:
                errors.append((i, f"Row {i}: Employee ID is required"))
                continue
            records.append({
                "employee_id":        eid,
                "full_name":          (row.get("Full Name")                   or row.get("full_name")          or "").strip() or None,
                "pan":                (row.get("PAN")                         or row.get("pan")                or "").strip() or None,
                "designation":        (row.get("Designation")                 or row.get("designation")        or "").strip() or None,
                "date_of_joining":    _safe_date(row.get("Date of Joining (YYYY-MM-DD)") or row.get("date_of_joining")),
                "date_of_birth":      _safe_date(row.get("Date of Birth (YYYY-MM-DD)")   or row.get("date_of_birth")),
                "appointment_nature": (row.get("Appointment Nature")          or row.get("appointment_nature") or "").strip() or None,
                "qualification":      (row.get("Qualification")               or row.get("qualification")      or "").strip() or None,
                "department":         (row.get("Department")                  or row.get("department")         or "").strip() or None,
                "experience_years":   _safe_int(row.get("Experience Years")   or row.get("experience_years")),
                "mobile":             (row.get("Mobile")                      or row.get("mobile")             or "").strip() or None,
                "email_personal":     (row.get("Personal Email")              or row.get("email_personal")     or "").strip() or None,
                "email_university":   (row.get("University Email")            or row.get("email_university")   or "").strip() or None,
                "specialization":     (row.get("Specialization")              or row.get("specialization")     or "").strip() or None,
            })
    else:
        raw = _read_workbook_rows(file_bytes)
        for i, row in enumerate(raw, start=3):
            eid = str(row[0]).strip() if row[0] else ""
            if not eid:
                errors.append((i, f"Row {i}: Employee ID is required"))
                continue
            records.append({
                "employee_id":        eid,
                "full_name":          str(row[1]).strip()  if row[1]  else None,
                "pan":                str(row[2]).strip()  if row[2]  else None,
                "designation":        str(row[3]).strip()  if row[3]  else None,
                "date_of_joining":    _safe_date(row[4]),
                "date_of_birth":      _safe_date(row[5]),
                "appointment_nature": str(row[6]).strip()  if row[6]  else None,
                "qualification":      str(row[7]).strip()  if row[7]  else None,
                "department":         str(row[8]).strip()  if row[8]  else None,
                "experience_years":   _safe_int(row[9]),
                "mobile":             str(row[10]).strip() if row[10] else None,
                "email_personal":     str(row[11]).strip() if row[11] else None,
                "email_university":   str(row[12]).strip() if row[12] else None,
                "specialization":     str(row[13]).strip() if row[13] else None,
            })
    return records, errors


# ============================================================
# BUILD BOOK CHAPTERS TEMPLATE
# ============================================================

BOOK_COLS = [
    ("Title of the Books Published *",                     50),
    ("Title of the Chapters Published",                    50),
    ("Category (Book / Chapter)",                          24),
    ("Title of Books Translated into Bhartiya Bhasha",     50),
    ("Title of Conference Proceedings / Publisher",         44),
    ("Category (Internal / External)",                     24),
    ("National / International",                           22),
    ("Date of Publication (YYYY-MM-DD)",                   24),
    ("ISBN Number",                                        22),
    ("Co-authors (if any)",                                36),
    ("DOI (if available)",                                 30),
    ("Indexed in Databases",                               30),
    ("Link to Journal Source",                             40),
    ("Supporting Doc Link",                                40),
]

BOOK_SAMPLE = [
    "Advances in Machine Learning",
    "Chapter 5: Deep Learning Techniques",
    "Chapter",
    "",
    "Springer Nature",
    "External",
    "International",
    "2024-06-15",
    "978-3-16-148410-0",
    "Dr. A. Kumar, Dr. B. Singh",
    "10.1007/978-3-030-12345-6",
    "Scopus, Web of Science",
    "https://link.springer.com/chapter/10.1007/978-3-030-12345-6",
    "https://drive.google.com/supporting-doc",
]

BOOK_TYPE_OPTIONS = ["Book", "Chapter"]
BOOK_IE_OPTIONS   = ["Internal", "External"]
BOOK_NI_OPTIONS   = ["National", "International"]


def build_book_chapters_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Book Chapters"
    ws.freeze_panes = "A3"
    n    = len(BOOK_COLS)
    last = get_column_letter(n)

    ws.merge_cells(f"A1:{last}1")
    t        = ws["A1"]
    t.value  = "Faculty MIS — Bulk Book / Chapter Upload Template"
    t.font   = Font(bold=True, size=12, color="0F172A")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 26

    _style_header(ws, 2, BOOK_COLS)
    _style_sample(ws, 3, BOOK_SAMPLE)

    ws.merge_cells(f"A4:{last}4")
    note        = ws["A4"]
    note.value  = "⚠  Delete the sample row (row 3) before uploading. * = required. Dates must be YYYY-MM-DD."
    note.font   = Font(size=8, color="94A3B8", italic=True)
    note.alignment = LEFT

    _add_dropdown(ws, "C", BOOK_TYPE_OPTIONS, start_row=5)
    _add_dropdown(ws, "F", BOOK_IE_OPTIONS,   start_row=5)
    _add_dropdown(ws, "G", BOOK_NI_OPTIONS,   start_row=5)

    for r in range(5, 1001):
        ws.row_dimensions[r].height = 16
        for c in range(1, n + 1):
            ws.cell(r, c).font   = BODY_FONT
            ws.cell(r, c).border = BORDER

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# PARSE BOOK CHAPTERS
# ============================================================

def parse_book_chapters(file_bytes: bytes, filename: str):
    """Returns (records, errors)."""
    records, errors = [], []
    is_csv = filename.lower().endswith(".csv")

    if is_csv:
        raw = _read_csv_rows(file_bytes)
        for i, row in enumerate(raw, start=2):
            book_title = (row.get("Title of the Books Published *") or
                          row.get("Title of the Books Published") or "").strip()
            if not book_title:
                errors.append((i, "Title of the Books Published is required"))
                continue
            records.append({
                "book_title":             book_title,
                "chapter_title":          (row.get("Title of the Chapters Published") or "").strip() or None,
                "book_or_chapter":        (row.get("Category (Book / Chapter)") or "").strip() or None,
                "translated_title":       (row.get("Title of Books Translated into Bhartiya Bhasha") or "").strip() or None,
                "proceedings_publisher":  (row.get("Title of Conference Proceedings / Publisher") or "").strip() or None,
                "internal_external":      (row.get("Category (Internal / External)") or "").strip() or None,
                "national_international": (row.get("National / International") or "").strip() or None,
                "publication_date":       _safe_date(row.get("Date of Publication (YYYY-MM-DD)") or row.get("Date of Publication")),
                "isbn":                   (row.get("ISBN Number") or "").strip() or None,
                "co_authors":             (row.get("Co-authors (if any)") or "").strip() or None,
                "doi":                    (row.get("DOI (if available)") or "").strip() or None,
                "indexed_in":             (row.get("Indexed in Databases") or "").strip() or None,
                "journal_link":           (row.get("Link to Journal Source") or "").strip() or None,
                "supporting_doc_link":    (row.get("Supporting Doc Link") or "").strip() or None,
            })
    else:
        raw = _read_workbook_rows(file_bytes)
        for i, row in enumerate(raw, start=3):
            book_title = str(row[0]).strip() if row[0] else ""
            if not book_title:
                errors.append((i, "Title of the Books Published is required"))
                continue
            records.append({
                "book_title":             book_title,
                "chapter_title":          str(row[1]).strip()  if row[1]  else None,
                "book_or_chapter":        str(row[2]).strip()  if row[2]  else None,
                "translated_title":       str(row[3]).strip()  if row[3]  else None,
                "proceedings_publisher":  str(row[4]).strip()  if row[4]  else None,
                "internal_external":      str(row[5]).strip()  if row[5]  else None,
                "national_international": str(row[6]).strip()  if row[6]  else None,
                "publication_date":       _safe_date(row[7]),
                "isbn":                   str(row[8]).strip()  if row[8]  else None,
                "co_authors":             str(row[9]).strip()  if row[9]  else None,
                "doi":                    str(row[10]).strip() if row[10] else None,
                "indexed_in":             str(row[11]).strip() if row[11] else None,
                "journal_link":           str(row[12]).strip() if row[12] else None,
                "supporting_doc_link":    str(row[13]).strip() if row[13] else None,
            })

    return records, errors


# ============================================================
# GUEST LECTURES — template + parser
# ============================================================

GUEST_COLS = [
    ("Lecture Title / Topic Delivered *",                  50),
    ("Organization & Location",                            40),
    ("Date (YYYY-MM-DD)",                                  22),
    ("JAIN / Outside",                                     18),
    ("Mode (In-person / Online)",                          22),
    ("Audience Type (Students / Faculty / Industry)",      28),
    ("Link to Brochure / Weblink",                         40),
    ("Link to Supporting Document",                        40),
]

GUEST_SAMPLE = [
    "Introduction to AI in Healthcare",
    "IIT Bangalore, Karnataka",
    "2024-08-15",
    "Outside",
    "In-person",
    "Students",
    "https://example.com/brochure",
    "https://drive.google.com/invite-letter",
]

GUEST_JAIN_OPTIONS    = ["JAIN", "Outside"]
GUEST_MODE_OPTIONS    = ["In-person", "Online"]
GUEST_AUDIENCE_OPTIONS = ["Students", "Faculty", "Industry"]


def build_guest_lectures_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Guest Lectures"
    ws.freeze_panes = "A3"
    n    = len(GUEST_COLS)
    last = get_column_letter(n)

    ws.merge_cells(f"A1:{last}1")
    t        = ws["A1"]
    t.value  = "Faculty MIS — Bulk Guest Lecture Upload Template"
    t.font   = Font(bold=True, size=12, color="0F172A")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 26

    _style_header(ws, 2, GUEST_COLS)
    _style_sample(ws, 3, GUEST_SAMPLE)

    ws.merge_cells(f"A4:{last}4")
    note        = ws["A4"]
    note.value  = "⚠  Delete the sample row (row 3) before uploading. * = required. Dates must be YYYY-MM-DD."
    note.font   = Font(size=8, color="94A3B8", italic=True)
    note.alignment = LEFT

    _add_dropdown(ws, "D", GUEST_JAIN_OPTIONS,     start_row=5)   # JAIN/Outside
    _add_dropdown(ws, "E", GUEST_MODE_OPTIONS,      start_row=5)   # Mode
    _add_dropdown(ws, "F", GUEST_AUDIENCE_OPTIONS,  start_row=5)   # Audience Type

    for r in range(5, 1001):
        ws.row_dimensions[r].height = 16
        for c in range(1, n + 1):
            ws.cell(r, c).font   = BODY_FONT
            ws.cell(r, c).border = BORDER

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# PARSE GUEST LECTURES
# ============================================================

def parse_guest_lectures(file_bytes: bytes, filename: str):
    """Returns (records, errors)."""
    records, errors = [], []
    is_csv = filename.lower().endswith(".csv")

    if is_csv:
        raw = _read_csv_rows(file_bytes)
        for i, row in enumerate(raw, start=2):
            title = (row.get("Lecture Title / Topic Delivered *") or
                     row.get("Lecture Title / Topic Delivered") or "").strip()
            if not title:
                errors.append((i, "Lecture Title / Topic Delivered is required"))
                continue
            records.append({
                "lecture_title":         title,
                "organization_location": (row.get("Organization & Location") or "").strip() or None,
                "lecture_date":          _safe_date(row.get("Date (YYYY-MM-DD)") or row.get("Date")),
                "jain_or_outside":       (row.get("JAIN / Outside") or "").strip() or None,
                "mode":                  (row.get("Mode (In-person / Online)") or row.get("Mode") or "").strip() or None,
                "audience_type":         (row.get("Audience Type (Students / Faculty / Industry)") or row.get("Audience Type") or "").strip() or None,
                "brochure_link":         (row.get("Link to Brochure / Weblink") or "").strip() or None,
                "supporting_doc_link":   (row.get("Link to Supporting Document") or "").strip() or None,
            })
    else:
        raw = _read_workbook_rows(file_bytes)
        for i, row in enumerate(raw, start=3):
            title = str(row[0]).strip() if row[0] else ""
            if not title:
                errors.append((i, "Lecture Title / Topic Delivered is required"))
                continue
            records.append({
                "lecture_title":         title,
                "organization_location": str(row[1]).strip() if row[1] else None,
                "lecture_date":          _safe_date(row[2]),
                "jain_or_outside":       str(row[3]).strip() if row[3] else None,
                "mode":                  str(row[4]).strip() if row[4] else None,
                "audience_type":         str(row[5]).strip() if row[5] else None,
                "brochure_link":         str(row[6]).strip() if row[6] else None,
                "supporting_doc_link":   str(row[7]).strip() if row[7] else None,
            })

    return records, errors


# ============================================================
# PATENTS / IP — template + parser
# ============================================================

PATENT_COLS = [
    ("Patent / Application Number",                         30),
    ("Type (Patent/Copyright/Trademark/GI/Design Reg.) *",  36),
    ("Title of Invention / Work *",                         50),
    ("Status (Filed/Published/Awarded)",                    24),
    ("Filing Date (YYYY-MM-DD)",                            22),
    ("Published Date (YYYY-MM-DD)",                         22),
    ("Grant Date (YYYY-MM-DD)",                             22),
    ("Awarding Agency",                                     34),
    ("National / International",                            22),
    ("Commercialization Details",                           40),
    ("Contribution to OER",                                 40),
    ("Link to Supporting Document",                         40),
]

PATENT_SAMPLE = [
    "IN202341012345",
    "Patent",
    "AI-based Diagnostic System for Cancer Detection",
    "Filed",
    "2024-03-15",
    "",
    "",
    "Indian Patent Office",
    "National",
    "",
    "",
    "https://drive.google.com/certificate",
]

PATENT_TYPE_OPTIONS   = ["Patent", "Copyright", "Trademark", "GI", "Design Registration"]
PATENT_STATUS_OPTIONS = ["Filed", "Published", "Awarded"]
PATENT_NI_OPTIONS     = ["National", "International"]


def build_patents_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patents"
    ws.freeze_panes = "A3"
    n    = len(PATENT_COLS)
    last = get_column_letter(n)

    ws.merge_cells(f"A1:{last}1")
    t        = ws["A1"]
    t.value  = "Faculty MIS — Bulk Patent / IP Upload Template"
    t.font   = Font(bold=True, size=12, color="0F172A")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 26

    _style_header(ws, 2, PATENT_COLS)
    _style_sample(ws, 3, PATENT_SAMPLE)

    ws.merge_cells(f"A4:{last}4")
    note        = ws["A4"]
    note.value  = "⚠  Delete the sample row (row 3) before uploading. * = required. Dates must be YYYY-MM-DD."
    note.font   = Font(size=8, color="94A3B8", italic=True)
    note.alignment = LEFT

    _add_dropdown(ws, "B", PATENT_TYPE_OPTIONS,   start_row=5)   # Type
    _add_dropdown(ws, "D", PATENT_STATUS_OPTIONS,  start_row=5)   # Status
    _add_dropdown(ws, "I", PATENT_NI_OPTIONS,      start_row=5)   # National/International

    for r in range(5, 1001):
        ws.row_dimensions[r].height = 16
        for c in range(1, n + 1):
            ws.cell(r, c).font   = BODY_FONT
            ws.cell(r, c).border = BORDER

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# PARSE PATENTS
# ============================================================

def parse_patents(file_bytes: bytes, filename: str):
    """Returns (records, errors)."""
    records, errors = [], []
    is_csv = filename.lower().endswith(".csv")

    if is_csv:
        raw = _read_csv_rows(file_bytes)
        for i, row in enumerate(raw, start=2):
            title = (row.get("Title of Invention / Work *") or
                     row.get("Title of Invention / Work") or "").strip()
            if not title:
                errors.append((i, "Title of Invention / Work is required"))
                continue
            records.append({
                "application_number":       (row.get("Patent / Application Number") or "").strip() or None,
                "ip_type":                  (row.get("Type (Patent/Copyright/Trademark/GI/Design Reg.) *") or
                                             row.get("Type") or "").strip() or None,
                "title":                    title,
                "status":                   (row.get("Status (Filed/Published/Awarded)") or
                                             row.get("Status") or "").strip() or None,
                "filing_date":              _safe_date(row.get("Filing Date (YYYY-MM-DD)") or row.get("Filing Date")),
                "published_date":           _safe_date(row.get("Published Date (YYYY-MM-DD)") or row.get("Published Date")),
                "grant_date":               _safe_date(row.get("Grant Date (YYYY-MM-DD)") or row.get("Grant Date")),
                "awarding_agency":          (row.get("Awarding Agency") or "").strip() or None,
                "national_international":   (row.get("National / International") or "").strip() or None,
                "commercialization_details":(row.get("Commercialization Details") or "").strip() or None,
                "oer_contribution":         (row.get("Contribution to OER") or "").strip() or None,
                "supporting_doc_link":      (row.get("Link to Supporting Document") or "").strip() or None,
            })
    else:
        raw = _read_workbook_rows(file_bytes)
        for i, row in enumerate(raw, start=3):
            title = str(row[2]).strip() if row[2] else ""
            if not title:
                errors.append((i, "Title of Invention / Work is required"))
                continue
            records.append({
                "application_number":       str(row[0]).strip()  if row[0]  else None,
                "ip_type":                  str(row[1]).strip()  if row[1]  else None,
                "title":                    title,
                "status":                   str(row[3]).strip()  if row[3]  else None,
                "filing_date":              _safe_date(row[4]),
                "published_date":           _safe_date(row[5]),
                "grant_date":               _safe_date(row[6]),
                "awarding_agency":          str(row[7]).strip()  if row[7]  else None,
                "national_international":   str(row[8]).strip()  if row[8]  else None,
                "commercialization_details":str(row[9]).strip()  if row[9]  else None,
                "oer_contribution":         str(row[10]).strip() if row[10] else None,
                "supporting_doc_link":      str(row[11]).strip() if row[11] else None,
            })

    return records, errors


# ============================================================
# FELLOWSHIPS — template + parser
# ============================================================

FELLOWSHIP_COLS = [
    ("Name of Award / Fellowship *",                       44),
    ("Financial Support (INR)",                            22),
    ("Purpose of the Grant",                               40),
    ("Type of Support",                                    24),
    ("National / International",                           22),
    ("Date of Award (YYYY-MM-DD)",                         22),
    ("Name of Awarding Agency",                            34),
    ("Duration of Fellowship",                             22),
    ("Research Topic / Area",                              40),
    ("Location of Study / Research",                       34),
    ("Collaborating Institution / Partner",                34),
    ("Link to Grant Letter",                               40),
]

FELLOWSHIP_SAMPLE = [
    "UGC Junior Research Fellowship",
    "5,00,000",
    "Doctoral Research in Machine Learning",
    "Research Grant",
    "National",
    "2024-06-15",
    "University Grants Commission",
    "2 Years",
    "Natural Language Processing",
    "IISc Bangalore",
    "MIT, USA",
    "https://drive.google.com/grant-letter",
]

FELLOWSHIP_NI_OPTIONS = ["National", "International"]


def build_fellowships_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fellowships"
    ws.freeze_panes = "A3"
    n    = len(FELLOWSHIP_COLS)
    last = get_column_letter(n)

    ws.merge_cells(f"A1:{last}1")
    t        = ws["A1"]
    t.value  = "Faculty MIS — Bulk Fellowship Upload Template"
    t.font   = Font(bold=True, size=12, color="0F172A")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 26

    _style_header(ws, 2, FELLOWSHIP_COLS)
    _style_sample(ws, 3, FELLOWSHIP_SAMPLE)

    ws.merge_cells(f"A4:{last}4")
    note        = ws["A4"]
    note.value  = "⚠  Delete the sample row (row 3) before uploading. * = required. Dates must be YYYY-MM-DD."
    note.font   = Font(size=8, color="94A3B8", italic=True)
    note.alignment = LEFT

    _add_dropdown(ws, "E", FELLOWSHIP_NI_OPTIONS, start_row=5)   # National/International

    for r in range(5, 1001):
        ws.row_dimensions[r].height = 16
        for c in range(1, n + 1):
            ws.cell(r, c).font   = BODY_FONT
            ws.cell(r, c).border = BORDER

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# PARSE FELLOWSHIPS
# ============================================================

def parse_fellowships(file_bytes: bytes, filename: str):
    """Returns (records, errors)."""
    records, errors = [], []
    is_csv = filename.lower().endswith(".csv")

    if is_csv:
        raw = _read_csv_rows(file_bytes)
        for i, row in enumerate(raw, start=2):
            name = (row.get("Name of Award / Fellowship *") or
                    row.get("Name of Award / Fellowship") or "").strip()
            if not name:
                errors.append((i, "Name of Award / Fellowship is required"))
                continue
            records.append({
                "award_name":               name,
                "financial_support":        (row.get("Financial Support (INR)") or "").strip() or None,
                "grant_purpose":            (row.get("Purpose of the Grant") or "").strip() or None,
                "support_type":             (row.get("Type of Support") or "").strip() or None,
                "national_international":   (row.get("National / International") or "").strip() or None,
                "award_date":               _safe_date(row.get("Date of Award (YYYY-MM-DD)") or row.get("Date of Award")),
                "awarding_agency":          (row.get("Name of Awarding Agency") or "").strip() or None,
                "duration":                 (row.get("Duration of Fellowship") or "").strip() or None,
                "research_topic":           (row.get("Research Topic / Area") or "").strip() or None,
                "location":                 (row.get("Location of Study / Research") or "").strip() or None,
                "collaborating_institution":(row.get("Collaborating Institution / Partner") or "").strip() or None,
                "grant_letter_link":        (row.get("Link to Grant Letter") or "").strip() or None,
            })
    else:
        raw = _read_workbook_rows(file_bytes)
        for i, row in enumerate(raw, start=3):
            name = str(row[0]).strip() if row[0] else ""
            if not name:
                errors.append((i, "Name of Award / Fellowship is required"))
                continue
            records.append({
                "award_name":               name,
                "financial_support":        str(row[1]).strip()  if row[1]  else None,
                "grant_purpose":            str(row[2]).strip()  if row[2]  else None,
                "support_type":             str(row[3]).strip()  if row[3]  else None,
                "national_international":   str(row[4]).strip()  if row[4]  else None,
                "award_date":               _safe_date(row[5]),
                "awarding_agency":          str(row[6]).strip()  if row[6]  else None,
                "duration":                 str(row[7]).strip()  if row[7]  else None,
                "research_topic":           str(row[8]).strip()  if row[8]  else None,
                "location":                 str(row[9]).strip()  if row[9]  else None,
                "collaborating_institution":str(row[10]).strip() if row[10] else None,
                "grant_letter_link":        str(row[11]).strip() if row[11] else None,
            })

    return records, errors


# =============================================================
# CONFERENCES PARTICIPATED — bulk template + parser
# =============================================================

CPART_COLS = [
    ("Conference Title *",                         40),
    ("Paper Title",                                36),
    ("Organisers",                                 30),
    ("Leads to Journal Publication (Yes/No)",      20),
    ("Paper Link",                                 36),
    ("Collaboration",                              30),
    ("Focus Area",                                 28),
    ("Date From (DD-MM-YYYY)",                     20),
    ("Date To (DD-MM-YYYY)",                       20),
    ("National/International",                     18),
    ("Proceedings Indexed (Yes/No)",               18),
    ("Indexing Details",                            36),
    ("Certificate Link",                           36),
]

CPART_SAMPLE = [
    "International Conference on AI & ML",
    "Deep Learning for NLP",
    "IEEE / IIT Madras",
    "Yes",
    "https://doi.org/10.1234/example",
    "Joint work with IISc Bangalore",
    "Research Methodology",
    "15-01-2025",
    "17-01-2025",
    "International",
    "Yes",
    "Scopus indexed, ISSN: 1234-5678",
    "https://drive.google.com/certificate",
]

CPART_NI_OPTIONS   = ["National", "International"]
CPART_YN_OPTIONS   = ["Yes", "No"]
CPART_FOCUS_OPTIONS = [
    "Research Methodology",
    "IPR",
    "Entrepreneurship",
    "Skill Development",
    "IKS",
]


def build_conferences_participated_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conferences Participated"
    ws.freeze_panes = "A3"
    n    = len(CPART_COLS)
    last = get_column_letter(n)

    ws.merge_cells(f"A1:{last}1")
    t        = ws["A1"]
    t.value  = "Faculty MIS — Bulk Conferences Participated Upload Template"
    t.font   = Font(bold=True, size=12, color="0F172A")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 26

    _style_header(ws, 2, CPART_COLS)
    _style_sample(ws, 3, CPART_SAMPLE)

    ws.merge_cells(f"A4:{last}4")
    note        = ws["A4"]
    note.value  = "⚠  Delete the sample row (row 3) before uploading. * = required. Dates must be DD-MM-YYYY."
    note.font   = Font(size=8, color="94A3B8", italic=True)
    note.alignment = LEFT

    # Dropdowns
    _add_dropdown(ws, "D", CPART_YN_OPTIONS, start_row=5)       # Leads to Pub
    _add_dropdown(ws, "G", CPART_FOCUS_OPTIONS, start_row=5)    # Focus Area
    _add_dropdown(ws, "J", CPART_NI_OPTIONS, start_row=5)       # Nat/Int
    _add_dropdown(ws, "K", CPART_YN_OPTIONS, start_row=5)       # Proceedings Indexed

    for r in range(5, 1001):
        ws.row_dimensions[r].height = 16
        for c in range(1, n + 1):
            ws.cell(r, c).font   = BODY_FONT
            ws.cell(r, c).border = BORDER

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_conferences_participated(file_bytes: bytes, filename: str):
    """Returns (records, errors)."""
    records, errors = [], []
    is_csv = filename.lower().endswith(".csv")

    if is_csv:
        raw = _read_csv_rows(file_bytes)
        for i, row in enumerate(raw, start=2):
            title = (row.get("Conference Title *") or row.get("Conference Title") or "").strip()
            if not title:
                errors.append((i, "Conference Title is required"))
                continue
            records.append({
                "conference_title":     title,
                "paper_title":          (row.get("Paper Title") or "").strip() or None,
                "organisers":           (row.get("Organisers") or "").strip() or None,
                "leads_to_publication": (row.get("Leads to Journal Publication (Yes/No)") or row.get("Leads to Publication") or "").strip() or None,
                "paper_link":           (row.get("Paper Link") or "").strip() or None,
                "collaboration":        (row.get("Collaboration") or "").strip() or None,
                "focus_area":           (row.get("Focus Area") or "").strip() or None,
                "date_from":            _safe_date(row.get("Date From (DD-MM-YYYY)") or row.get("Date From")),
                "date_to":              _safe_date(row.get("Date To (DD-MM-YYYY)") or row.get("Date To")),
                "national_international": (row.get("National/International") or row.get("Nat/Int") or "").strip() or None,
                "proceedings_indexed":  (row.get("Proceedings Indexed (Yes/No)") or row.get("Proceedings Indexed") or "").strip() or None,
                "indexing_details":     (row.get("Indexing Details") or "").strip() or None,
                "certificate_link":     (row.get("Certificate Link") or "").strip() or None,
            })
    else:
        raw = _read_workbook_rows(file_bytes)
        for i, row in enumerate(raw, start=3):
            title = str(row[0]).strip() if row[0] else ""
            if not title:
                errors.append((i, "Conference Title is required"))
                continue
            records.append({
                "conference_title":      title,
                "paper_title":           str(row[1]).strip()  if row[1]  else None,
                "organisers":            str(row[2]).strip()  if row[2]  else None,
                "leads_to_publication":  str(row[3]).strip()  if row[3]  else None,
                "paper_link":            str(row[4]).strip()  if row[4]  else None,
                "collaboration":         str(row[5]).strip()  if row[5]  else None,
                "focus_area":            str(row[6]).strip()  if row[6]  else None,
                "date_from":             _safe_date(row[7]),
                "date_to":               _safe_date(row[8]),
                "national_international":str(row[9]).strip()  if row[9]  else None,
                "proceedings_indexed":   str(row[10]).strip() if row[10] else None,
                "indexing_details":      str(row[11]).strip() if row[11] else None,
                "certificate_link":      str(row[12]).strip() if row[12] else None,
            })

    return records, errors


# =============================================================
# CONFERENCES ORGANISED — bulk template + parser
# =============================================================

CORG_COLS = [
    ("Title *",                                    40),
    ("Department",                                 26),
    ("Role of the Faculty",                        26),
    ("Collaboration",                              30),
    ("Focus Area",                                 28),
    ("National/International",                     18),
    ("Number of Participants",                     16),
    ("Date From (DD-MM-YYYY)",                     20),
    ("Date To (DD-MM-YYYY)",                       20),
    ("Proceedings Indexed (Yes/No)",               18),
    ("Indexing Details",                            36),
    ("Activity Report Link",                       36),
]

CORG_SAMPLE = [
    "National Workshop on Cyber Security",
    "Computer Science",
    "Convener",
    "In collaboration with CDAC",
    "Skill Development",
    "National",
    150,
    "10-01-2025",
    "12-01-2025",
    "No",
    "",
    "https://drive.google.com/report",
]

CORG_NI_OPTIONS   = ["National", "International"]
CORG_YN_OPTIONS   = ["Yes", "No"]
CORG_FOCUS_OPTIONS = [
    "Research Methodology",
    "IPR",
    "Entrepreneurship",
    "Skill Development",
    "IKS",
]


def build_conferences_organised_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conferences Organised"
    ws.freeze_panes = "A3"
    n    = len(CORG_COLS)
    last = get_column_letter(n)

    ws.merge_cells(f"A1:{last}1")
    t        = ws["A1"]
    t.value  = "Faculty MIS — Bulk Conferences Organised Upload Template"
    t.font   = Font(bold=True, size=12, color="0F172A")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 26

    _style_header(ws, 2, CORG_COLS)
    _style_sample(ws, 3, CORG_SAMPLE)

    ws.merge_cells(f"A4:{last}4")
    note        = ws["A4"]
    note.value  = "⚠  Delete the sample row (row 3) before uploading. * = required. Dates must be DD-MM-YYYY."
    note.font   = Font(size=8, color="94A3B8", italic=True)
    note.alignment = LEFT

    # Dropdowns
    _add_dropdown(ws, "E", CORG_FOCUS_OPTIONS, start_row=5)    # Focus Area
    _add_dropdown(ws, "F", CORG_NI_OPTIONS, start_row=5)       # Nat/Int
    _add_dropdown(ws, "J", CORG_YN_OPTIONS, start_row=5)       # Proceedings Indexed

    for r in range(5, 1001):
        ws.row_dimensions[r].height = 16
        for c in range(1, n + 1):
            ws.cell(r, c).font   = BODY_FONT
            ws.cell(r, c).border = BORDER

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_conferences_organised(file_bytes: bytes, filename: str):
    """Returns (records, errors)."""
    records, errors = [], []
    is_csv = filename.lower().endswith(".csv")

    if is_csv:
        raw = _read_csv_rows(file_bytes)
        for i, row in enumerate(raw, start=2):
            title = (row.get("Title *") or row.get("Title") or "").strip()
            if not title:
                errors.append((i, "Title is required"))
                continue
            records.append({
                "title":                title,
                "department":           (row.get("Department") or "").strip() or None,
                "faculty_role":         (row.get("Role of the Faculty") or row.get("Faculty Role") or "").strip() or None,
                "collaboration":        (row.get("Collaboration") or "").strip() or None,
                "focus_area":           (row.get("Focus Area") or "").strip() or None,
                "national_international": (row.get("National/International") or row.get("Nat/Int") or "").strip() or None,
                "num_participants":     _safe_int(row.get("Number of Participants") or row.get("No. of Participants")),
                "date_from":            _safe_date(row.get("Date From (DD-MM-YYYY)") or row.get("Date From")),
                "date_to":              _safe_date(row.get("Date To (DD-MM-YYYY)") or row.get("Date To")),
                "proceedings_indexed":  (row.get("Proceedings Indexed (Yes/No)") or row.get("Proceedings Indexed") or "").strip() or None,
                "indexing_details":     (row.get("Indexing Details") or "").strip() or None,
                "activity_report_link": (row.get("Activity Report Link") or "").strip() or None,
            })
    else:
        raw = _read_workbook_rows(file_bytes)
        for i, row in enumerate(raw, start=3):
            title = str(row[0]).strip() if row[0] else ""
            if not title:
                errors.append((i, "Title is required"))
                continue
            records.append({
                "title":                 title,
                "department":            str(row[1]).strip()  if row[1]  else None,
                "faculty_role":          str(row[2]).strip()  if row[2]  else None,
                "collaboration":         str(row[3]).strip()  if row[3]  else None,
                "focus_area":            str(row[4]).strip()  if row[4]  else None,
                "national_international":str(row[5]).strip()  if row[5]  else None,
                "num_participants":      _safe_int(row[6]),
                "date_from":             _safe_date(row[7]),
                "date_to":               _safe_date(row[8]),
                "proceedings_indexed":   str(row[9]).strip()  if row[9]  else None,
                "indexing_details":      str(row[10]).strip() if row[10] else None,
                "activity_report_link":  str(row[11]).strip() if row[11] else None,
            })

    return records, errors


# =============================================================
# FDP PARTICIPATED — bulk template + parser
# =============================================================

FPART_COLS = [
    ("Program Title *",                            40),
    ("Duration (Days)",                            14),
    ("Start Date (DD-MM-YYYY)",                    20),
    ("End Date (DD-MM-YYYY)",                      20),
    ("Program Type",                               24),
    ("National/International",                     18),
    ("Organizing/Sponsoring Agency",               34),
    ("Location",                                   30),
    ("Mode (Online/Offline/Hybrid)",               18),
    ("Funding (Self/Institution)",                 18),
    ("Certificate Link",                           36),
    ("Brochure Link",                              36),
    ("Enrolled in Coursera/Online Platform (Yes/No)", 24),
]

FPART_SAMPLE = [
    "FDP on Machine Learning & AI",
    "5",
    "15-01-2025",
    "19-01-2025",
    "FDP",
    "National",
    "AICTE",
    "IIT Bombay, Mumbai",
    "Offline",
    "Institution",
    "https://drive.google.com/cert",
    "https://drive.google.com/brochure",
    "No",
]

FPART_NI_OPTIONS    = ["National", "International"]
FPART_MODE_OPTIONS  = ["Online", "Offline", "Hybrid"]
FPART_FUND_OPTIONS  = ["Self", "Institution"]
FPART_YN_OPTIONS    = ["Yes", "No"]


def build_fdp_participated_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FDP Participated"
    ws.freeze_panes = "A3"
    n    = len(FPART_COLS)
    last = get_column_letter(n)

    ws.merge_cells(f"A1:{last}1")
    t        = ws["A1"]
    t.value  = "Faculty MIS — Bulk FDP Participated Upload Template"
    t.font   = Font(bold=True, size=12, color="0F172A")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 26

    _style_header(ws, 2, FPART_COLS)
    _style_sample(ws, 3, FPART_SAMPLE)

    ws.merge_cells(f"A4:{last}4")
    note        = ws["A4"]
    note.value  = "⚠  Delete the sample row (row 3) before uploading. * = required. Dates must be DD-MM-YYYY."
    note.font   = Font(size=8, color="94A3B8", italic=True)
    note.alignment = LEFT

    # Dropdowns
    _add_dropdown(ws, "F", FPART_NI_OPTIONS, start_row=5)       # Nat/Int
    _add_dropdown(ws, "I", FPART_MODE_OPTIONS, start_row=5)     # Mode
    _add_dropdown(ws, "J", FPART_FUND_OPTIONS, start_row=5)     # Funding
    _add_dropdown(ws, "M", FPART_YN_OPTIONS, start_row=5)       # Enrolled Coursera

    for r in range(5, 1001):
        ws.row_dimensions[r].height = 16
        for c in range(1, n + 1):
            ws.cell(r, c).font   = BODY_FONT
            ws.cell(r, c).border = BORDER

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_fdp_participated(file_bytes: bytes, filename: str):
    """Returns (records, errors)."""
    records, errors = [], []
    is_csv = filename.lower().endswith(".csv")

    if is_csv:
        raw = _read_csv_rows(file_bytes)
        for i, row in enumerate(raw, start=2):
            title = (row.get("Program Title *") or row.get("Program Title") or "").strip()
            if not title:
                errors.append((i, "Program Title is required"))
                continue
            records.append({
                "program_title":        title,
                "duration_days":        (row.get("Duration (Days)") or row.get("Duration") or "").strip() or None,
                "start_date":           _safe_date(row.get("Start Date (DD-MM-YYYY)") or row.get("Start Date")),
                "end_date":             _safe_date(row.get("End Date (DD-MM-YYYY)") or row.get("End Date")),
                "program_type":         (row.get("Program Type") or "").strip() or None,
                "national_international": (row.get("National/International") or row.get("Nat/Int") or "").strip() or None,
                "organizing_agency":    (row.get("Organizing/Sponsoring Agency") or row.get("Organizing Agency") or "").strip() or None,
                "location":             (row.get("Location") or "").strip() or None,
                "mode":                 (row.get("Mode (Online/Offline/Hybrid)") or row.get("Mode") or "").strip() or None,
                "funding":              (row.get("Funding (Self/Institution)") or row.get("Funding") or "").strip() or None,
                "certificate_link":     (row.get("Certificate Link") or "").strip() or None,
                "brochure_link":        (row.get("Brochure Link") or "").strip() or None,
                "enrolled_coursera":    (row.get("Enrolled in Coursera/Online Platform (Yes/No)") or row.get("Enrolled in Coursera") or "").strip() or None,
            })
    else:
        raw = _read_workbook_rows(file_bytes)
        for i, row in enumerate(raw, start=3):
            title = str(row[0]).strip() if row[0] else ""
            if not title:
                errors.append((i, "Program Title is required"))
                continue
            records.append({
                "program_title":         title,
                "duration_days":         str(row[1]).strip()  if row[1]  else None,
                "start_date":            _safe_date(row[2]),
                "end_date":              _safe_date(row[3]),
                "program_type":          str(row[4]).strip()  if row[4]  else None,
                "national_international":str(row[5]).strip()  if row[5]  else None,
                "organizing_agency":     str(row[6]).strip()  if row[6]  else None,
                "location":              str(row[7]).strip()  if row[7]  else None,
                "mode":                  str(row[8]).strip()  if row[8]  else None,
                "funding":               str(row[9]).strip()  if row[9]  else None,
                "certificate_link":      str(row[10]).strip() if row[10] else None,
                "brochure_link":         str(row[11]).strip() if row[11] else None,
                "enrolled_coursera":     str(row[12]).strip() if row[12] else None,
            })

    return records, errors


# =============================================================
# FDP ORGANISED — bulk template + parser
# =============================================================

FORG_COLS = [
    ("Title *",                                    40),
    ("Department",                                 26),
    ("Role of the Faculty",                        26),
    ("Collaboration",                              30),
    ("Focus Area",                                 28),
    ("National/International",                     18),
    ("Number of Participants",                     16),
    ("Date From (DD-MM-YYYY)",                     20),
    ("Date To (DD-MM-YYYY)",                       20),
    ("Activity Report Link",                       36),
]

FORG_SAMPLE = [
    "FDP on Advanced Data Analytics",
    "Computer Science",
    "Convener",
    "In collaboration with AICTE",
    "Skill Development",
    "National",
    50,
    "10-02-2025",
    "14-02-2025",
    "https://drive.google.com/report",
]

FORG_NI_OPTIONS    = ["National", "International"]
FORG_FOCUS_OPTIONS = [
    "Research Methodology",
    "IPR",
    "Entrepreneurship",
    "Skill Development",
    "IKS",
]


def build_fdp_organised_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FDP Organised"
    ws.freeze_panes = "A3"
    n    = len(FORG_COLS)
    last = get_column_letter(n)

    ws.merge_cells(f"A1:{last}1")
    t        = ws["A1"]
    t.value  = "Faculty MIS — Bulk FDP Organised Upload Template"
    t.font   = Font(bold=True, size=12, color="0F172A")
    t.alignment = CENTER
    ws.row_dimensions[1].height = 26

    _style_header(ws, 2, FORG_COLS)
    _style_sample(ws, 3, FORG_SAMPLE)

    ws.merge_cells(f"A4:{last}4")
    note        = ws["A4"]
    note.value  = "⚠  Delete the sample row (row 3) before uploading. * = required. Dates must be DD-MM-YYYY."
    note.font   = Font(size=8, color="94A3B8", italic=True)
    note.alignment = LEFT

    # Dropdowns
    _add_dropdown(ws, "E", FORG_FOCUS_OPTIONS, start_row=5)    # Focus Area
    _add_dropdown(ws, "F", FORG_NI_OPTIONS, start_row=5)       # Nat/Int

    for r in range(5, 1001):
        ws.row_dimensions[r].height = 16
        for c in range(1, n + 1):
            ws.cell(r, c).font   = BODY_FONT
            ws.cell(r, c).border = BORDER

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_fdp_organised(file_bytes: bytes, filename: str):
    """Returns (records, errors)."""
    records, errors = [], []
    is_csv = filename.lower().endswith(".csv")

    if is_csv:
        raw = _read_csv_rows(file_bytes)
        for i, row in enumerate(raw, start=2):
            title = (row.get("Title *") or row.get("Title") or "").strip()
            if not title:
                errors.append((i, "Title is required"))
                continue
            records.append({
                "title":                title,
                "department":           (row.get("Department") or "").strip() or None,
                "faculty_role":         (row.get("Role of the Faculty") or row.get("Faculty Role") or "").strip() or None,
                "collaboration":        (row.get("Collaboration") or "").strip() or None,
                "focus_area":           (row.get("Focus Area") or "").strip() or None,
                "national_international": (row.get("National/International") or row.get("Nat/Int") or "").strip() or None,
                "num_participants":     _safe_int(row.get("Number of Participants") or row.get("No. of Participants")),
                "date_from":            _safe_date(row.get("Date From (DD-MM-YYYY)") or row.get("Date From")),
                "date_to":              _safe_date(row.get("Date To (DD-MM-YYYY)") or row.get("Date To")),
                "activity_report_link": (row.get("Activity Report Link") or "").strip() or None,
            })
    else:
        raw = _read_workbook_rows(file_bytes)
        for i, row in enumerate(raw, start=3):
            title = str(row[0]).strip() if row[0] else ""
            if not title:
                errors.append((i, "Title is required"))
                continue
            records.append({
                "title":                 title,
                "department":            str(row[1]).strip()  if row[1]  else None,
                "faculty_role":          str(row[2]).strip()  if row[2]  else None,
                "collaboration":         str(row[3]).strip()  if row[3]  else None,
                "focus_area":            str(row[4]).strip()  if row[4]  else None,
                "national_international":str(row[5]).strip()  if row[5]  else None,
                "num_participants":      _safe_int(row[6]),
                "date_from":             _safe_date(row[7]),
                "date_to":              _safe_date(row[8]),
                "activity_report_link":  str(row[9]).strip()  if row[9]  else None,
            })

    return records, errors
